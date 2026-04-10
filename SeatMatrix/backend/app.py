"""
SeatMatrix - Smart Exam Hall Seating Arrangement System
Sasurie College of Engineering (Autonomous)
Version: 2.0 — Production Ready
"""

from flask import Flask, render_template, request, jsonify, send_file, session, make_response
from flask_cors import CORS
import json, re, csv, io, os, random, hashlib, uuid, sqlite3
from datetime import datetime
from functools import wraps
from dotenv import load_dotenv
from openai import OpenAI

try:
    import psycopg2
    from psycopg2.extras import DictCursor
    POSTGRES_SUPPORTED = True
except ImportError:
    POSTGRES_SUPPORTED = False


try:
    import openpyxl
    EXCEL_SUPPORTED = True
except ImportError:
    EXCEL_SUPPORTED = False

try:
    import fitz
    PDF_PARSE_SUPPORTED = True
except ImportError:
    PDF_PARSE_SUPPORTED = False

load_dotenv()

app = Flask(__name__, 
            template_folder='../frontend/templates', 
            static_folder='../frontend/static')
app.secret_key = os.getenv("SECRET_KEY", "seatmatrix-secret-change-me")
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "database", "seatmatrix.db")
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)


DATABASE_URL = os.getenv("DATABASE_URL")

class DBWrapper:
    def __init__(self, conn, is_pg):
        self.conn = conn
        self.is_pg = is_pg
    def execute(self, query, params=None):
        if self.is_pg:
            query = query.replace('?', '%s')
            cur = self.conn.cursor(cursor_factory=DictCursor)
            cur.execute(query, params or ())
            return cur
        return self.conn.execute(query, params or ())
    def commit(self): self.conn.commit()
    def close(self): self.conn.close()
    def cursor(self):
        if self.is_pg: return self.conn.cursor(cursor_factory=DictCursor)
        return self.conn.cursor()


def get_db():
    if DATABASE_URL:
        # PostgreSQL connection
        conn = psycopg2.connect(DATABASE_URL)
        return DBWrapper(conn, True)
    else:
        # SQLite connection
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return DBWrapper(conn, False)


def init_db():
    conn = get_db()
    c = conn.cursor()
    script = """
    CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY, password TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'staff', name TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS exams (
        id TEXT PRIMARY KEY, name TEXT NOT NULL, subject TEXT,
        subject_code TEXT, date TEXT, time TEXT, duration TEXT,
        paper_sets TEXT DEFAULT '2', department TEXT, semester TEXT,
        year_subject_map TEXT DEFAULT '{}', year_subject_code_map TEXT DEFAULT '{}', created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS rooms (
        id TEXT PRIMARY KEY, name TEXT NOT NULL, block TEXT,
        capacity INTEGER, grid_rows INTEGER DEFAULT 5, grid_cols INTEGER DEFAULT 8,
        blocked_seats TEXT DEFAULT '[]', location TEXT, aisle_after_col TEXT DEFAULT '[]',
        college_id TEXT DEFAULT 'sasurie', created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS staff (
        id TEXT PRIMARY KEY, name TEXT NOT NULL, email TEXT,
        department TEXT, phone TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS students (
        id TEXT PRIMARY KEY, student_name TEXT NOT NULL,
        register_number TEXT UNIQUE NOT NULL, department TEXT,
        year TEXT, subject TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS colleges (
        id TEXT PRIMARY KEY, name TEXT NOT NULL, address TEXT,
        city TEXT, notes TEXT, created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS college_halls (
        id TEXT PRIMARY KEY, college_id TEXT NOT NULL, name TEXT NOT NULL,
        block TEXT, location TEXT, capacity INTEGER,
        grid_rows INTEGER DEFAULT 5, grid_cols INTEGER DEFAULT 8,
        blocked_seats TEXT DEFAULT '[]', aisle_after_col TEXT DEFAULT '[]',
        notes TEXT, created_at TEXT,
        FOREIGN KEY (college_id) REFERENCES colleges(id)
    );
    CREATE TABLE IF NOT EXISTS seating (
        id TEXT PRIMARY KEY, exam_id TEXT NOT NULL,
        data TEXT NOT NULL, created_at TEXT
    );
    """
    if DATABASE_URL:
        # PostgreSQL doesn't have executescript, run as one or split
        c.execute(script)
    else:
        # sqlite3 executescript handles multiple statements
        c.executescript(script)


    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        c.execute("INSERT INTO users VALUES (?,?,?,?)", ("admin", hashlib.sha256("admin123".encode()).hexdigest(), "admin", "Administrator"))
        c.execute("INSERT INTO users VALUES (?,?,?,?)", ("staff1", hashlib.sha256("staff123".encode()).hexdigest(), "staff", "Dr. Priya Kumar"))

    if c.execute("SELECT COUNT(*) FROM colleges WHERE id='sasurie'").fetchone()[0] == 0:
        c.execute("INSERT INTO colleges VALUES (?,?,?,?,?,?)",
            ("sasurie","Sasurie College of Engineering",
             "Vijayamangalam, Tiruppur District, Tamil Nadu - 638056",
             "Vijayamangalam","Autonomous Institution",datetime.now().isoformat()))
        halls = [
            ("hall-101","sasurie","Hall 101","Main Block","Ground Floor",40,5,8,"[]",None,""),
            ("hall-102","sasurie","Hall 102","Main Block","Ground Floor",40,5,8,"[]",None,""),
            ("hall-201","sasurie","Hall 201","Main Block","First Floor", 48,6,8,"[]",None,""),
            ("hall-202","sasurie","Hall 202","Main Block","First Floor", 48,6,8,"[]",None,""),
            ("lab-cs1", "sasurie","CS Lab 1", "CS Block", "Ground Floor",30,5,6,"[]",None,""),
            ("lab-cs2", "sasurie","CS Lab 2", "CS Block", "First Floor", 30,5,6,"[]",None,""),
            ("seminar", "sasurie","Seminar Hall","Admin Block","Ground Floor",100,10,10,"[]",5,"Large hall"),
            ("hall-301","sasurie","Hall 301","Main Block","Second Floor",40,5,8,"[]",None,""),
        ]
        for h in halls:
            c.execute("INSERT INTO college_halls VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", h+(datetime.now().isoformat(),))

    conn.commit()
    # Migration: add new columns to existing databases
    try: conn.execute("ALTER TABLE exams ADD COLUMN year_subject_map TEXT DEFAULT '{}'"); conn.commit()
    except: pass
    try: conn.execute("ALTER TABLE exams ADD COLUMN year_subject_code_map TEXT DEFAULT '{}'"); conn.commit()
    except: pass
    # Migrate aisle_after_col from int to JSON array
    try:
        rows = conn.execute("SELECT id, aisle_after_col FROM rooms").fetchall()
        for row in rows:
            val = row[1]
            if val and not str(val).startswith('['):
                try:
                    new_val = json.dumps([int(val)])
                    conn.execute("UPDATE rooms SET aisle_after_col=? WHERE id=?", (new_val, row[0]))
                except: pass
        rows2 = conn.execute("SELECT id, aisle_after_col FROM college_halls").fetchall()
        for row in rows2:
            val = row[1]
            if val and not str(val).startswith('['):
                try:
                    new_val = json.dumps([int(val)])
                    conn.execute("UPDATE college_halls SET aisle_after_col=? WHERE id=?", (new_val, row[0]))
                except: pass
        conn.commit()
    except: pass
    conn.close()


init_db()

def row_to_dict(row): return dict(row) if row else None
def rows_to_list(rows): return [dict(r) for r in rows]

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session: return jsonify({"error":"Unauthorized"}),401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session: return jsonify({"error":"Unauthorized"}),401
        conn = get_db()
        u = row_to_dict(conn.execute("SELECT role FROM users WHERE username=?",(session["user"],)).fetchone())
        conn.close()
        if not u or u["role"] != "admin": return jsonify({"error":"Admin access required"}),403
        return f(*args, **kwargs)
    return decorated

# ── PAGES ────────────────────────────────────────────────────────────────────
@app.route("/")
def index(): return render_template("index.html")

@app.route("/favicon.ico")
def favicon():
    return send_file("static/images/sasurie_invert_logo-removebg-preview.png", mimetype="image/png")

# ── AUTH ─────────────────────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def login():
    d = request.json or {}
    pw = hashlib.sha256(d.get("password","").encode()).hexdigest()
    conn = get_db()
    user = row_to_dict(conn.execute("SELECT * FROM users WHERE username=?",(d.get("username",""),)).fetchone())
    conn.close()
    if user and user["password"] == pw:
        session["user"] = user["username"]
        session["role"] = user["role"]
        session["name"] = user["name"]
        return jsonify({"success":True,"role":user["role"],"name":user["name"]})
    return jsonify({"error":"Invalid username or password"}),401

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear(); return jsonify({"success":True})

@app.route("/api/me")
def me():
    if "user" not in session: return jsonify({"error":"Not logged in","logged_in":False}),401
    conn = get_db()
    user = row_to_dict(conn.execute("SELECT username,role,name FROM users WHERE username=?",(session["user"],)).fetchone())
    conn.close()
    if not user: session.clear(); return jsonify({"error":"User not found","logged_in":False}),401
    return jsonify({"logged_in":True,"username":user["username"],"role":user["role"],"name":user["name"]})

# ── STATS ─────────────────────────────────────────────────────────────────────
@app.route("/api/stats")
@login_required
def stats():
    conn = get_db()
    seating_rows = conn.execute("SELECT data FROM seating").fetchall()
    seated_regnos = set()
    for row in seating_rows:
        try:
            arr = json.loads(row[0])
            for s in arr.get("seats",[]):
                if s.get("register_number"): seated_regnos.add(s["register_number"])
        except: pass
    data = {
        "total_exams":    conn.execute("SELECT COUNT(*) FROM exams").fetchone()[0],
        "total_rooms":    conn.execute("SELECT COUNT(*) FROM rooms").fetchone()[0],
        "total_staff":    conn.execute("SELECT COUNT(*) FROM staff").fetchone()[0],
        "total_students": conn.execute("SELECT COUNT(*) FROM students").fetchone()[0],
        "total_colleges": conn.execute("SELECT COUNT(*) FROM colleges").fetchone()[0],
        "total_seating":  conn.execute("SELECT COUNT(*) FROM seating").fetchone()[0],
        "unique_seated":  len(seated_regnos),
        "total_users":    conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
    }
    conn.close(); return jsonify(data)

# ── EXAMS ─────────────────────────────────────────────────────────────────────
@app.route("/api/exams", methods=["GET","POST"])
@login_required
def exams():
    if request.method == "GET":
        conn = get_db()
        exams = rows_to_list(conn.execute("SELECT * FROM exams ORDER BY date DESC, created_at DESC").fetchall())
        conn.close()
        for e in exams:
            raw = e.get("year_subject_map","{}")
            try: e["year_subject_map"] = json.loads(raw) if isinstance(raw,str) else raw
            except: e["year_subject_map"] = {}
            raw_code = e.get("year_subject_code_map","{}")
            try: e["year_subject_code_map"] = json.loads(raw_code) if isinstance(raw_code,str) else raw_code
            except: e["year_subject_code_map"] = {}
        return jsonify(exams)
    d = request.json or {}
    if not d.get("name"): return jsonify({"error":"Exam name required"}),400
    e = {"id":str(uuid.uuid4()),"name":d["name"],"subject":d.get("subject",""),
         "subject_code":d.get("subject_code",""),"date":d.get("date",""),
         "time":d.get("time",""),"duration":d.get("duration",""),
         "paper_sets":d.get("paper_sets","2"),"department":d.get("department",""),
         "semester":d.get("semester",""),
         "year_subject_map":json.dumps(d.get("year_subject_map",{})),
         "year_subject_code_map":json.dumps(d.get("year_subject_code_map",{})),
         "created_at":datetime.now().isoformat()}
    conn = get_db()
    conn.execute("INSERT INTO exams VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",tuple(e.values()))
    conn.commit(); conn.close()
    e["year_subject_map"] = d.get("year_subject_map",{})
    e["year_subject_code_map"] = d.get("year_subject_code_map", {})
    return jsonify(e), 201

@app.route("/api/exams/<eid>", methods=["PUT","DELETE"])
@login_required
def exam_detail(eid):
    conn = get_db()
    if request.method == "DELETE":
        conn.execute("DELETE FROM exams WHERE id=?",(eid,))
        conn.execute("DELETE FROM seating WHERE exam_id=?",(eid,))
        conn.commit(); conn.close(); return jsonify({"success":True})
    d = request.json or {}
    conn.execute(
        "UPDATE exams SET name=?,subject=?,subject_code=?,date=?,time=?,duration=?,paper_sets=?,department=?,semester=?,year_subject_map=?,year_subject_code_map=? WHERE id=?",
        (d.get("name"),d.get("subject",""),d.get("subject_code",""),d.get("date",""),
         d.get("time",""),d.get("duration",""),d.get("paper_sets","2"),
         d.get("department",""),d.get("semester",""),json.dumps(d.get("year_subject_map",{})),json.dumps(d.get("year_subject_code_map",{})),eid))
    conn.commit(); conn.close(); return jsonify({"success":True})

# ── ROOMS ─────────────────────────────────────────────────────────────────────
@app.route("/api/rooms", methods=["GET","POST"])
@login_required
def rooms():
    if request.method == "GET":
        conn = get_db()
        rooms = rows_to_list(conn.execute("SELECT * FROM rooms ORDER BY name").fetchall())
        conn.close()
        for r in rooms: r["blocked_seats"] = json.loads(r.get("blocked_seats") or "[]")
        return jsonify(rooms)
    d = request.json or {}
    if not d.get("name"): return jsonify({"error":"Room name required"}),400
    blocked = d.get("blocked_seats",[])
    r = {"id":str(uuid.uuid4()),"name":d["name"],"block":d.get("block",""),
         "capacity":int(d.get("capacity",40)),"grid_rows":int(d.get("grid_rows",5)),
         "grid_cols":int(d.get("grid_cols",8)),
         "blocked_seats":json.dumps(blocked if isinstance(blocked,list) else []),
         "location":d.get("location",""),
         "aisle_after_col":json.dumps([int(x) for x in str(d.get("aisle_after_col","")).replace(" ","").split(",") if x.strip().isdigit()] if d.get("aisle_after_col") else []),
         "college_id":d.get("college_id","sasurie"),"created_at":datetime.now().isoformat()}
    conn = get_db()
    conn.execute("INSERT INTO rooms VALUES (?,?,?,?,?,?,?,?,?,?,?)",tuple(r.values()))
    conn.commit(); conn.close()
    r["blocked_seats"] = blocked; return jsonify(r), 201

@app.route("/api/rooms/<rid>", methods=["PUT","DELETE"])
@login_required
def room_detail(rid):
    conn = get_db()
    if request.method == "DELETE":
        conn.execute("DELETE FROM rooms WHERE id=?",(rid,))
        conn.commit(); conn.close(); return jsonify({"success":True})
    d = request.json or {}
    conn.execute(
        "UPDATE rooms SET name=?,block=?,capacity=?,grid_rows=?,grid_cols=?,blocked_seats=?,location=?,aisle_after_col=? WHERE id=?",
        (d.get("name"),d.get("block"),d.get("capacity"),d.get("grid_rows"),d.get("grid_cols"),
         json.dumps(d.get("blocked_seats",[])),d.get("location"),d.get("aisle_after_col"),rid))
    conn.commit(); conn.close(); return jsonify({"success":True})

# ── STAFF ─────────────────────────────────────────────────────────────────────
@app.route("/api/staff", methods=["GET","POST"])
@login_required
def staff():
    if request.method == "GET":
        conn = get_db()
        s = rows_to_list(conn.execute("SELECT * FROM staff ORDER BY name").fetchall())
        conn.close(); return jsonify(s)
    d = request.json or {}
    if not d.get("name"): return jsonify({"error":"Name required"}),400
    s = {"id":str(uuid.uuid4()),"name":d["name"],"email":d.get("email",""),
         "department":d.get("department",""),"phone":d.get("phone",""),
         "created_at":datetime.now().isoformat()}
    conn = get_db()
    conn.execute("INSERT INTO staff VALUES (?,?,?,?,?,?)",tuple(s.values()))
    conn.commit(); conn.close(); return jsonify(s), 201

@app.route("/api/staff/<sid>", methods=["PUT","DELETE"])
@login_required
def staff_detail(sid):
    conn = get_db()
    if request.method == "DELETE":
        conn.execute("DELETE FROM staff WHERE id=?",(sid,))
        conn.commit(); conn.close(); return jsonify({"success":True})
    d = request.json or {}
    conn.execute("UPDATE staff SET name=?,email=?,department=?,phone=? WHERE id=?",
        (d.get("name"),d.get("email",""),d.get("department",""),d.get("phone",""),sid))
    conn.commit(); conn.close(); return jsonify({"success":True})

# ── STUDENTS ──────────────────────────────────────────────────────────────────
@app.route("/api/students", methods=["GET","POST"])
@login_required
def students():
    if request.method == "GET":
        conn = get_db()
        q = request.args.get("q",""); dept = request.args.get("dept",""); yr = request.args.get("year","")
        query = "SELECT * FROM students WHERE 1=1"
        params = []
        if q: query += " AND (student_name LIKE ? OR register_number LIKE ?)"; params += [f"%{q}%",f"%{q}%"]
        if dept: query += " AND department=?"; params.append(dept)
        if yr: query += " AND year=?"; params.append(yr)
        query += " ORDER BY register_number"
        s = rows_to_list(conn.execute(query,params).fetchall())
        conn.close(); return jsonify(s)
    d = request.json or {}
    if not d.get("register_number"): return jsonify({"error":"Register number required"}),400
    auto_dept, auto_year = decode_register_number(d.get("register_number",""))
    s = {"id":str(uuid.uuid4()),
         "student_name":d.get("student_name","—"),
         "register_number":d["register_number"],
         "department":d.get("department","") or auto_dept,
         "year":str(d.get("year","")) or auto_year,
         "subject":d.get("subject",""),
         "created_at":datetime.now().isoformat()}
    try:
        conn = get_db()
        conn.execute("INSERT INTO students VALUES (?,?,?,?,?,?,?)",tuple(s.values()))
        conn.commit(); conn.close(); return jsonify(s), 201
    except sqlite3.IntegrityError:
        return jsonify({"error":"Register number already exists"}),400

@app.route("/api/students/<sid>", methods=["PUT","DELETE"])
@login_required
def student_detail(sid):
    conn = get_db()
    if request.method == "DELETE":
        conn.execute("DELETE FROM students WHERE id=?",(sid,))
        conn.commit(); conn.close(); return jsonify({"success":True})
    d = request.json or {}
    conn.execute(
        "UPDATE students SET student_name=?,register_number=?,department=?,year=?,subject=? WHERE id=?",
        (d.get("student_name"),d.get("register_number"),
         d.get("department",""),d.get("year",""),d.get("subject",""),sid))
    conn.commit(); conn.close(); return jsonify({"success":True})

# ── USERS ─────────────────────────────────────────────────────────────────────
@app.route("/api/users", methods=["GET","POST"])
@admin_required
def users():
    if request.method == "GET":
        conn = get_db()
        u = rows_to_list(conn.execute("SELECT username,role,name FROM users ORDER BY name").fetchall())
        conn.close(); return jsonify(u)
    d = request.json or {}
    username = (d.get("username") or "").strip()
    name     = (d.get("name") or "").strip()
    password = d.get("password","")
    role     = d.get("role","staff")
    if not username or not name or not password:
        return jsonify({"error":"Username, full name and password are all required"}),400
    if len(password) < 6:
        return jsonify({"error":"Password must be at least 6 characters"}),400
    if role not in ("admin","staff"):
        return jsonify({"error":"Role must be admin or staff"}),400
    conn = get_db()
    try:
        conn.execute("INSERT INTO users VALUES (?,?,?,?)",
            (username, hashlib.sha256(password.encode()).hexdigest(), role, name))
        conn.commit(); conn.close()
        return jsonify({"success":True,"username":username,"name":name,"role":role}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error":f"Username \'{username}\' is already taken. Choose another."}),400

@app.route("/api/users/<username>", methods=["DELETE"])
@admin_required
def delete_user(username):
    if username == "admin": return jsonify({"error":"The main admin account cannot be deleted"}),400
    if username == session.get("user"): return jsonify({"error":"You cannot delete your own account"}),400
    conn = get_db()
    conn.execute("DELETE FROM users WHERE username=?",(username,))
    conn.commit(); conn.close()
    return jsonify({"success":True})

@app.route("/api/users/<username>/password", methods=["PUT"])
@login_required
def change_password(username):
    caller_role = session.get("role","staff")
    if caller_role != "admin" and session.get("user") != username:
        return jsonify({"error":"You can only change your own password"}),403
    d = request.json or {}
    password = d.get("password","")
    if len(password) < 6: return jsonify({"error":"Password must be at least 6 characters"}),400
    conn = get_db()
    conn.execute("UPDATE users SET password=? WHERE username=?",
        (hashlib.sha256(password.encode()).hexdigest(), username))
    conn.commit(); conn.close()
    return jsonify({"success":True})

# ─────────────────────────────────────────────
# REGISTER NUMBER DECODER (Sasurie format: 73 + YY + YY + DDD + NNN)
# ─────────────────────────────────────────────
DEPT_MAP = {
    '243': 'AIDS',    '205': 'IT',     '104': 'CSE',    '105': 'ECE',
    '106': 'EEE',     '103': 'MECH',   '114': 'CIVIL',  '149': 'CSE CS',
    '631': 'MBA',     '107': 'EIE',    '108': 'ICE',    '149': 'CSE CS',
    '106': 'EEE',     '301': 'Special'
}
YEAR_MAP = {'21': '4', '22': '3', '23': '3', '24': '2', '25': '1'}

def decode_register_number(regno):
    """Auto-decode department and year from Sasurie register number format."""
    regno = str(regno).strip()
    if len(regno) == 12 and regno.startswith('73'):
        dept_code = regno[6:9]
        adm_yr    = regno[4:6]
        dept = DEPT_MAP.get(dept_code, '')
        year = YEAR_MAP.get(adm_yr, '')
        return dept, year
    return '', ''


@app.route("/api/students/bulk", methods=["POST"])
@login_required
def bulk_import_students():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename.lower()
    added = skipped = 0
    rows_to_insert = []

    # ── CSV ──────────────────────────────────────────
    if filename.endswith(".csv"):
        try:
            content = file.read().decode("utf-8-sig")
        except UnicodeDecodeError:
            content = file.read().decode("latin-1")

        # Handle plain list of register numbers (single column, no header)
        lines = [l.strip() for l in content.strip().splitlines() if l.strip()]
        is_plain_list = all(re.match(r'^73\d{10}$', l.split(',')[0].strip()) for l in lines[:5] if l)

        if is_plain_list:
            for line in lines:
                regno = line.split(',')[0].strip()
                if not regno: continue
                dept, year = decode_register_number(regno)
                rows_to_insert.append({
                    "student_name": "—", "register_number": regno,
                    "department": dept, "year": year, "subject": ""
                })
        else:
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                regno = (row.get("Register Number") or row.get("register_number") or
                         row.get("RegNo") or row.get("regno") or "").strip()
                if not regno: skipped += 1; continue
                # Auto-decode if dept/year missing
                auto_dept, auto_year = decode_register_number(regno)
                rows_to_insert.append({
                    "student_name": (row.get("Name") or row.get("name") or row.get("Student Name") or "—").strip(),
                    "register_number": regno,
                    "department": (row.get("Department") or row.get("department") or row.get("Dept") or "").strip() or auto_dept,
                    "year": (row.get("Year") or row.get("year") or "").strip() or auto_year,
                    "subject": (row.get("Subject") or row.get("subject") or "").strip(),
                })

    # ── EXCEL (.xlsx / .xls) ─────────────────────────
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        if not EXCEL_SUPPORTED:
            return jsonify({"error": "Excel support not installed. Run: pip install openpyxl"}), 400
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h.lower(): i for i, h in enumerate(headers)}
        def gc(row, *keys):
            for k in keys:
                for hk, idx in col.items():
                    if k.lower() in hk:
                        v = row[idx].value
                        return str(v).strip() if v else ""
            return ""
        for row in ws.iter_rows(min_row=2):
            regno = gc(row, "register", "regno", "reg no", "roll")
            if not regno: skipped += 1; continue
            auto_dept, auto_year = decode_register_number(regno)
            rows_to_insert.append({
                "student_name": gc(row, "name", "student") or "—",
                "register_number": regno,
                "department": gc(row, "department", "dept") or auto_dept,
                "year": gc(row, "year") or auto_year,
                "subject": gc(row, "subject"),
            })

    # ── PDF — extract register numbers automatically ──
    elif filename.endswith(".pdf"):
        if not PDF_PARSE_SUPPORTED:
            return jsonify({"error": "PDF parsing not installed. Run: pip install PyMuPDF"}), 400
        raw_text = ""
        doc = fitz.open(stream=file.read(), filetype="pdf")
        for page in doc:
            raw_text += page.get_text()
        regnos = list(dict.fromkeys(re.findall(r'73\d{10}', raw_text)))
        for r in regnos:
            dept, year = decode_register_number(r)
            rows_to_insert.append({
                "student_name": "—", "register_number": r,
                "department": dept, "year": year, "subject": ""
            })
    else:
        return jsonify({"error": "Unsupported file type. Use CSV, XLSX, or PDF"}), 400

    # ── Insert into DB ──────────────────────────────
    conn = get_db()
    for s in rows_to_insert:
        if not s["register_number"]: skipped += 1; continue
        try:
            conn.execute("INSERT INTO students VALUES (?,?,?,?,?,?,?)", (
                str(uuid.uuid4()), s["student_name"], s["register_number"],
                s["department"], s["year"], s["subject"], datetime.now().isoformat()))
            added += 1
        except sqlite3.IntegrityError:
            skipped += 1
    conn.commit(); conn.close()
    return jsonify({"added": added, "skipped": skipped, "total": added + skipped,
                    "message": f"Imported {added} students. {skipped} skipped (duplicates or empty)."})



@app.route("/api/students/export")
@login_required
def export_students():
    conn = get_db()
    students = rows_to_list(conn.execute("SELECT * FROM students ORDER BY register_number").fetchall())
    conn.close()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Name","Register Number","Department","Year","Subject"])
    for s in students:
        w.writerow([s["student_name"],s["register_number"],s["department"],s["year"],s["subject"]])
    out.seek(0)
    return send_file(io.BytesIO(out.read().encode()), mimetype="text/csv",
                     as_attachment=True, download_name="students.csv")


# ── COLLEGES ─────────────────────────────────────────────────────────────────
@app.route("/api/colleges", methods=["GET","POST"])
@login_required
def colleges():
    if request.method == "GET":
        conn = get_db()
        cols = rows_to_list(conn.execute("SELECT * FROM colleges ORDER BY name").fetchall())
        for col in cols:
            halls = rows_to_list(conn.execute(
                "SELECT * FROM college_halls WHERE college_id=? ORDER BY name",(col["id"],)).fetchall())
            for h in halls: h["blocked_seats"] = json.loads(h.get("blocked_seats") or "[]")
            col["halls"] = halls
        conn.close(); return jsonify(cols)
    d = request.json or {}
    if not d.get("name"): return jsonify({"error":"College name required"}),400
    c = {"id":str(uuid.uuid4()),"name":d["name"],"address":d.get("address",""),
         "city":d.get("city",""),"notes":d.get("notes",""),"created_at":datetime.now().isoformat()}
    conn = get_db()
    conn.execute("INSERT INTO colleges VALUES (?,?,?,?,?,?)",tuple(c.values()))
    conn.commit(); conn.close(); return jsonify(c), 201

@app.route("/api/colleges/<cid>", methods=["DELETE"])
@login_required
def delete_college(cid):
    if cid == "sasurie": return jsonify({"error":"Cannot delete built-in Sasurie preset"}),400
    conn = get_db()
    conn.execute("DELETE FROM college_halls WHERE college_id=?",(cid,))
    conn.execute("DELETE FROM colleges WHERE id=?",(cid,))
    conn.commit(); conn.close(); return jsonify({"success":True})

@app.route("/api/colleges/<cid>/halls", methods=["POST"])
@login_required
def add_college_hall(cid):
    d = request.json or {}
    if not d.get("name"): return jsonify({"error":"Hall name required"}),400
    blocked = d.get("blocked_seats",[])
    h = {"id":str(uuid.uuid4()),"college_id":cid,"name":d["name"],
         "block":d.get("block",""),"location":d.get("location",""),
         "capacity":int(d.get("capacity",0)),
         "grid_rows":int(d.get("grid_rows",5)),"grid_cols":int(d.get("grid_cols",8)),
         "blocked_seats":json.dumps(blocked if isinstance(blocked,list) else []),
         "aisle_after_col":json.dumps([int(x) for x in str(d.get("aisle_after_col","")).replace(" ","").split(",") if x.strip().isdigit()] if d.get("aisle_after_col") else []),
         "notes":d.get("notes",""),"created_at":datetime.now().isoformat()}
    conn = get_db()
    conn.execute("INSERT INTO college_halls VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",tuple(h.values()))
    conn.commit(); conn.close()
    h["blocked_seats"] = blocked; return jsonify(h), 201

@app.route("/api/colleges/<cid>/halls/<hid>", methods=["DELETE"])
@login_required
def delete_college_hall(cid, hid):
    conn = get_db()
    conn.execute("DELETE FROM college_halls WHERE id=? AND college_id=?",(hid,cid))
    conn.commit(); conn.close(); return jsonify({"success":True})

@app.route("/api/colleges/<cid>/halls/<hid>/load", methods=["POST"])
@login_required
def load_hall_to_rooms(cid, hid):
    conn = get_db()
    h = row_to_dict(conn.execute("SELECT * FROM college_halls WHERE id=? AND college_id=?",(hid,cid)).fetchone())
    if not h: conn.close(); return jsonify({"error":"Hall not found"}),404
    rid = str(uuid.uuid4())
    conn.execute("INSERT INTO rooms VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (rid,h["name"],h["block"],h["capacity"],h["grid_rows"],h["grid_cols"],
         h["blocked_seats"],h["location"],h["aisle_after_col"],cid,datetime.now().isoformat()))
    conn.commit(); conn.close()
    return jsonify({"success":True,"room_id":rid,"message":"Hall added to active rooms"})


# ── SEATING ALGORITHM ───────────────────────────────────────────────────────
def anti_malpractice_seating(students, rooms, settings, exam=None):
    """
    Smart anti-malpractice seating:
    - Proportionally distributes each group (dept+year) across ALL halls
    - True checkerboard A/B paper sets — no two adjacent seats same set
    - Sorts students by register number within each group
    - Interleaves groups seat-by-seat: Y1-CSE, Y2-ECE, Y1-CSE, Y2-ECE...
    """
    import math
    n_sets = int(settings.get("paper_sets", 2))
    labels = ["A", "B", "C", "D"][:n_sets]

    # Build year_subject_map
    year_subject_map = {}
    default_subject = ""
    year_subject_code_map = {}
    if exam:
        raw = exam.get("year_subject_map", "{}")
        if isinstance(raw, str):
            try: year_subject_map = json.loads(raw)
            except: year_subject_map = {}
        elif isinstance(raw, dict):
            year_subject_map = raw
        default_subject = exam.get("subject", "") or ""
        raw_code = exam.get("year_subject_code_map", "{}")
        if isinstance(raw_code, str):
            try: year_subject_code_map = json.loads(raw_code)
            except: year_subject_code_map = {}
        elif isinstance(raw_code, dict):
            year_subject_code_map = raw_code

    def resolve_subject(stu, year):
        dept = str(stu.get("department") or stu.get("Department") or "").strip()
        yr = str(year).strip()
        for key in [f"{yr}|{dept}", f"{yr}|", f"|{dept}", yr, ""]:
            if key in year_subject_map and year_subject_map[key]:
                return year_subject_map[key]
        return (stu.get("subject") or stu.get("Subject") or default_subject or "")

    def resolve_subject_code(stu, year):
        dept = str(stu.get("department") or stu.get("Department") or "").strip()
        yr = str(year).strip()
        for key in [f"{yr}|{dept}", f"{yr}|", f"|{dept}", yr, ""]:
            if key in year_subject_code_map and year_subject_code_map[key]:
                return year_subject_code_map[key]
        return exam.get("subject_code", "") if exam else ""

    # ── Group students by dept+year ──────────────────────────────────────────
    groups = {}
    for s in students:
        dept = s.get("department") or s.get("Department") or "Unknown"
        year = str(s.get("year", "") or "")
        key = f"{dept}|{year}"
        groups.setdefault(key, []).append(s)

    # Sort each group by register number
    for key in groups:
        groups[key].sort(key=lambda s: str(s.get("register_number") or s.get("Register Number") or ""))

    group_keys = list(groups.keys())
    total_students = sum(len(g) for g in groups.values())

    # ── Calculate positions per room ─────────────────────────────────────────
    def parse_aisles(raw):
        if not raw: return []
        if isinstance(raw, list): return [int(x) for x in raw if str(x).strip().isdigit() or isinstance(x, int)]
        raw = str(raw).strip()
        if raw.startswith('['):
            try: return [int(x) for x in json.loads(raw)]
            except: return []
        return [int(x.strip()) for x in raw.split(',') if x.strip().isdigit()]

    room_positions = []  # list of (room, positions_list, aisles)
    total_capacity = 0
    for room in rooms:
        rows_r = room.get("grid_rows", 5)
        cols_r = room.get("grid_cols", 8)
        blocked = room.get("blocked_seats", [])
        if isinstance(blocked, str): blocked = json.loads(blocked)
        capacity = room.get("capacity", rows_r * cols_r)
        aisles = parse_aisles(room.get("aisle_after_col"))
        positions = [(r, c, f"R{r}C{c}") for r in range(1, rows_r + 1)
                     for c in range(1, cols_r + 1)
                     if f"R{r}C{c}" not in blocked][:capacity]
        room_positions.append((room, positions, aisles))
        total_capacity += len(positions)

    # ── PROPORTIONAL DISTRIBUTION across halls ────────────────────────────────
    # Each group gets seats proportionally in EACH hall
    # e.g. 50 Y1-CSE + 50 Y2-ECE across 2 halls of 60:
    #   Hall 1: 25 Y1-CSE + 25 Y2-ECE (interleaved)
    #   Hall 2: 25 Y1-CSE + 25 Y2-ECE (interleaved)

    # Allocate how many from each group go into each room
    students_capped = min(total_students, total_capacity)
    scale = students_capped / total_students if total_students > 0 else 1.0

    # ── ALLOCATE students to rooms using cumulative distribution ─────────────
    # This avoids rounding errors that drop students from small groups across multiple halls
    room_queues = []
    group_allocated_total = {k: 0 for k in group_keys}
    cumulative_capacity = 0

    for room_idx, (room, positions, aisles) in enumerate(room_positions):
        room_cap = len(positions)
        cumulative_capacity += room_cap
        
        # Fraction of total capacity utilized up to this room
        cum_fraction = cumulative_capacity / total_capacity if total_capacity > 0 else 0
        
        room_group_students = {}
        for gkey in group_keys:
            # Total students from this group that should be seated in all rooms processed so far
            group_total_to_seat = len(groups[gkey]) * scale
            target_cumulative = round(group_total_to_seat * cum_fraction)
            
            # Allocation for this specific room
            count = target_cumulative - group_allocated_total[gkey]
            
            # Ensure we don't accidentally seat more than available in this group
            count = max(0, min(count, len(groups[gkey]) - group_allocated_total[gkey]))
            
            room_group_students[gkey] = count
            group_allocated_total[gkey] += count

        # Normalization: Occasionally rounding might slightly exceed room capacity
        allocated_in_room = sum(room_group_students.values())
        if allocated_in_room > room_cap:
            diff = allocated_in_room - room_cap
            sorted_keys = sorted(group_keys, key=lambda k: room_group_students[k], reverse=True)
            for i in range(diff):
                k = sorted_keys[i % len(sorted_keys)]
                if room_group_students[k] > 0:
                    room_group_students[k] -= 1
                    group_allocated_total[k] -= 1

        room_queues.append(room_group_students)


    # ── Slice students into per-room allocations ──────────────────────────────
    group_offsets = {k: 0 for k in group_keys}
    room_student_lists = []
    for room_idx, room_group_students in enumerate(room_queues):
        room_groups = {}
        for gkey in group_keys:
            count = max(0, room_group_students.get(gkey, 0))
            start = group_offsets[gkey]
            end = start + count
            # Don't exceed available students
            end = min(end, len(groups[gkey]))
            count = end - start
            if count > 0:
                room_groups[gkey] = groups[gkey][start:end]
                group_offsets[gkey] = end
        room_student_lists.append(room_groups)

    # ── Interleave groups per room and assign seats ───────────────────────────
    all_seats = []
    seat_counter = 1

    for room_idx, (room, positions, aisles) in enumerate(room_positions):
        room_groups = room_student_lists[room_idx]
        if not room_groups:
            continue

        # True round-robin interleave of all groups in this room
        group_queues = {k: v[:] for k, v in room_groups.items()}
        ordered = []
        group_order = list(group_queues.keys())
        while any(group_queues.values()):
            for gkey in group_order:
                if group_queues.get(gkey):
                    ordered.append(group_queues[gkey].pop(0))

        # Assign seats
        for i, (r, c, label) in enumerate(positions):
            if i >= len(ordered):
                break
            stu = ordered[i]
            student_year = str(stu.get("year", "") or "").strip()
            subject = resolve_subject(stu, student_year)
            subject_code = resolve_subject_code(stu, student_year)
            all_seats.append({
                "seat_number": seat_counter,
                "seat_label": label, "row": r, "col": c,
                "student_name": stu.get("Name") or stu.get("student_name", ""),
                "register_number": stu.get("Register Number") or stu.get("register_number", ""),
                "department": stu.get("Department") or stu.get("department", ""),
                "year": student_year,
                "subject": subject,
                "subject_code": subject_code,
                "room_id": room["id"], "room_name": room["name"],
                "paper_set": labels[(r + c) % n_sets],
                "aisles": aisles,
                "grid_cols": room.get("grid_cols", 8),
            })
            seat_counter += 1

    return all_seats


@app.route("/api/generate-seating", methods=["POST"])
@login_required
def generate_seating():
    data = request.json
    conn = get_db()
    exam = row_to_dict(conn.execute("SELECT * FROM exams WHERE id=?",(data.get("exam_id"),)).fetchone())
    if not exam: conn.close(); return jsonify({"error":"Exam not found"}),404
    # Parse year_subject_map from JSON string
    raw_ysm = exam.get("year_subject_map", "{}")
    try: exam["year_subject_map"] = json.loads(raw_ysm) if isinstance(raw_ysm, str) else raw_ysm
    except: exam["year_subject_map"] = {}

    # Also accept year_subject_code_map from request (passed by frontend)
    exam["year_subject_code_map"] = data.get("year_subject_code_map", {})

    rooms = []
    invigilators = []
    
    # Handle append_mode: filter out rooms already in previous arrangement
    existing_room_ids = set()
    if data.get("append_mode"):
        existing_row = conn.execute("SELECT data FROM seating WHERE exam_id=? ORDER BY created_at DESC LIMIT 1", (data.get("exam_id"),)).fetchone()
        if existing_row:
            existing_data = json.loads(existing_row["data"])
            existing_room_ids = {r["id"] for r in existing_data.get("rooms", [])}

    for rid in data.get("room_ids",[]):
        if rid in existing_room_ids: continue
        r = row_to_dict(conn.execute("SELECT * FROM rooms WHERE id=?",(rid,)).fetchone())
        if r: r["blocked_seats"] = json.loads(r.get("blocked_seats") or "[]"); rooms.append(r)
    
    for iid in data.get("invigilator_ids",[]):
        s = row_to_dict(conn.execute("SELECT * FROM staff WHERE id=?",(iid,)).fetchone())
        if s: invigilators.append(s)

    if not rooms: conn.close(); return jsonify({"error":"No NEW rooms selected for generation" if data.get("append_mode") else "No valid rooms selected"}),400

    students = data.get("students",[])
    if not students: conn.close(); return jsonify({"error":"No students provided"}),400

    # Calculate total capacity across all selected rooms
    total_capacity = sum(
        len([(r2,c2) for r2 in range(1, rm.get("grid_rows",5)+1)
                     for c2 in range(1, rm.get("grid_cols",8)+1)
                     if f"R{r2}C{c2}" not in (rm.get("blocked_seats") or [])])
        for rm in rooms
    )
    total_capacity = min(total_capacity, sum(rm.get("capacity", 40) for rm in rooms))

    # Only seat as many students as the halls can hold
    students_to_seat = students[:total_capacity]
    remaining_students = students[total_capacity:]

    seats = anti_malpractice_seating(students_to_seat, rooms, data.get("settings",{}), exam=exam)

    # Assign invigilators to seats permanently
    room_ids_ordered = list(dict.fromkeys(rm["id"] for rm in rooms))
    for i, rid in enumerate(room_ids_ordered):
        inv_name = invigilators[i % len(invigilators)]["name"] if invigilators else "TBD"
        for s in seats:
            if s.get("room_id") == rid:
                s["invigilator"] = inv_name

    arrangement = {
        "exam_id": data.get("exam_id"),
        "exam": exam,
        "rooms": rooms,
        "invigilators": invigilators,
        "seats": seats,
        "settings": data.get("settings",{}),
        "generated_at": datetime.now().isoformat(),
        "total_students_provided": len(students),
        "total_seated": len(seats),
        "total_remaining": len(remaining_students),
    }

    # Store arrangement (append mode — keeps previous halls, adds new ones)
    existing_row = conn.execute(
        "SELECT * FROM seating WHERE exam_id=? ORDER BY created_at DESC LIMIT 1",
        (data.get("exam_id"),)).fetchone()

    if existing_row and data.get("append_mode"):
        # Merge with existing arrangement
        existing = json.loads(existing_row["data"])
        existing_seats = existing.get("seats", [])
        # Re-number seats from where we left off
        offset = len(existing_seats)
        for s in seats:
            s["seat_number"] += offset
        merged_seats = existing_seats + seats
        merged_rooms = existing.get("rooms", []) + [r for r in rooms if r["id"] not in {x["id"] for x in existing.get("rooms",[])}]
        arrangement["seats"] = merged_seats
        arrangement["rooms"] = merged_rooms
        arrangement["total_seated"] = len(merged_seats)
        existing_id = existing_row["id"]
        conn.execute("UPDATE seating SET data=?, created_at=? WHERE id=?",
            (json.dumps(arrangement), datetime.now().isoformat(), existing_id))
    else:
        sid = str(uuid.uuid4())
        conn.execute("INSERT OR REPLACE INTO seating VALUES (?,?,?,?)",
            (sid, data.get("exam_id"), json.dumps(arrangement), datetime.now().isoformat()))

    conn.commit(); conn.close()

    return jsonify({
        "success": True,
        "arrangement": arrangement,
        "total_seated": len(seats),
        "total_remaining": len(remaining_students),
        "remaining_students": remaining_students,  # send back so frontend can use for next hall
        "message": f"Seated {len(seats)} students in {len(rooms)} hall(s). {len(remaining_students)} students remaining."
    })


@app.route("/api/seating/<exam_id>")
@login_required
def get_seating(exam_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM seating WHERE exam_id=? ORDER BY created_at DESC LIMIT 1",(exam_id,)).fetchone()
    conn.close()
    if not row: return jsonify({"error":"No seating found"}),404
    return jsonify(json.loads(row["data"]))

@app.route("/api/archives")
@login_required
def get_archives():
    conn = get_db()
    rows = rows_to_list(conn.execute(
        "SELECT s.id,s.exam_id,s.data,s.created_at,e.name,e.subject,e.date,e.time,e.duration FROM seating s LEFT JOIN exams e ON s.exam_id=e.id ORDER BY s.created_at DESC"
    ).fetchall())
    conn.close()
    result = []
    for r in rows:
        try:
            arr = json.loads(r["data"])
            seats = arr.get("seats", [])
            total_students = len(seats)
            room_ids = list(dict.fromkeys(s.get("room_id") for s in seats))
            rooms_used = len(room_ids)
            room_names = list(dict.fromkeys(s.get("room_name","") for s in seats))
            # Dept/year breakdown per hall
            from collections import Counter as _C
            dept_year = sorted(set(f"{s.get('department','')} Y{s.get('year','')}" for s in seats))
            invigilators_list = arr.get("invigilators", [])
            halls_detail = []
            for i, rid in enumerate(room_ids):
                hall_seats = [s for s in seats if s.get("room_id")==rid]
                hall_name = hall_seats[0].get("room_name","") if hall_seats else ""
                hall_depts = sorted(set(f"{s.get('department','')} Y{s.get('year','')}" for s in hall_seats))
                invigilator_name = invigilators_list[i % len(invigilators_list)]["name"] if invigilators_list else "—"
                halls_detail.append({"id":rid,"name":hall_name,"students":len(hall_seats),"groups":hall_depts,"invigilator":invigilator_name})
        except:
            total_students=0; rooms_used=0; room_names=[]; dept_year=[]; halls_detail=[]
        # Format time
        raw_time = r["time"] or ""
        nice_time = ""
        if raw_time:
            try:
                h,m = raw_time.split(":"); h=int(h)
                nice_time = f"{h if h<=12 else h-12}:{m} {'AM' if h<12 else 'PM'}"
            except: nice_time = raw_time
        result.append({
            "id": r["id"], "exam_id": r["exam_id"],
            "exam_name": r["name"] or "Unknown",
            "subject": r["subject"] or "",
            "exam_date": r["date"] or "",
            "exam_time": nice_time,
            "duration": r["duration"] or "",
            "generated_at": r["created_at"],
            "total_students": total_students,
            "rooms_used": rooms_used,
            "room_names": room_names,
            "dept_year_groups": dept_year,
            "halls_detail": halls_detail,
        })
    return jsonify(result)

@app.route("/api/archives/<archive_id>", methods=["DELETE"])
@login_required
def delete_archive(archive_id):
    conn = get_db()
    conn.execute("DELETE FROM seating WHERE id=?", (archive_id,))
    conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route("/api/students/<sid>", methods=["DELETE"])
@login_required
def delete_student(sid):
    conn = get_db()
    conn.execute("DELETE FROM students WHERE id=?", (sid,))
    conn.commit(); conn.close(); return jsonify({"success":True})

@app.route("/api/students", methods=["DELETE"])
@login_required
def delete_all_students():
    conn = get_db()
    conn.execute("DELETE FROM students")
    conn.commit(); conn.close(); return jsonify({"success":True})

@app.route("/api/find-seat", methods=["POST"])
def find_seat():
    query = (request.json.get("query") or "").strip().upper()
    if not query:
        return jsonify({"error": "Please enter your register number"}), 400
    if len(query) < 4:
        return jsonify({"error": "Please enter a valid register number"}), 400

    conn = get_db()
    # Get only the LATEST seating arrangement per exam (avoid duplicates)
    rows = rows_to_list(conn.execute(
        "SELECT * FROM seating ORDER BY created_at DESC"
    ).fetchall())
    conn.close()

    # Deduplicate: keep only latest arrangement per exam_id
    seen_exams = set()
    unique_rows = []
    for row in rows:
        if row["exam_id"] not in seen_exams:
            seen_exams.add(row["exam_id"])
            unique_rows.append(row)

    results = []
    for row in unique_rows:
        try:
            arr = json.loads(row["data"])
        except: continue
        exam = arr.get("exam", {})

        # Build year_subject_map
        ysm_raw = exam.get("year_subject_map", "{}")
        try:
            ysm = json.loads(ysm_raw) if isinstance(ysm_raw, str) else (ysm_raw or {})
        except: ysm = {}

        yscm_raw = exam.get("year_subject_code_map", "{}")
        try:
            yscm = json.loads(yscm_raw) if isinstance(yscm_raw, str) else (yscm_raw or {})
        except: yscm = {}

        for seat in arr.get("seats", []):
            regno = (seat.get("register_number") or "").strip().upper()
            if regno != query:
                continue

            # Resolve subject using all key formats
            subject = seat.get("subject", "")
            if not subject:
                year = str(seat.get("year", "")).strip()
                dept = str(seat.get("department", "")).strip()
                for key in [f"{year}|{dept}", f"{year}|", f"|{dept}", year, ""]:
                    if key in ysm and ysm[key]:
                        subject = ysm[key]
                        break
            if not subject:
                subject = exam.get("subject", "")

            # Resolve subject_code
            subject_code = seat.get("subject_code", "")
            if not subject_code:
                year = str(seat.get("year", "")).strip()
                dept = str(seat.get("department", "")).strip()
                for key in [f"{year}|{dept}", f"{year}|", f"|{dept}", year, ""]:
                    if key in yscm and yscm[key]:
                        subject_code = yscm[key]
                        break
            if not subject_code:
                subject_code = exam.get("subject_code", "")

            # Format time nicely
            raw_time = exam.get("time", "")
            nice_time = ""
            if raw_time:
                try:
                    h, m = raw_time.split(":")
                    h = int(h)
                    nice_time = f"{h if h <= 12 else h-12}:{m} {'AM' if h < 12 else 'PM'}"
                except: nice_time = raw_time

            results.append({
                "student_name": seat.get("student_name") or "—",
                "register_number": seat.get("register_number"),
                "department": seat.get("department", ""),
                "year": seat.get("year", ""),
                "exam_name": exam.get("name", ""),
                "exam_date": exam.get("date", ""),
                "exam_time": nice_time,
                "duration": exam.get("duration", ""),
                "subject": subject,
                "subject_code": subject_code,
                "room": seat.get("room_name"),
                "seat_number": seat.get("seat_number"),
                "seat_label": seat.get("seat_label"),
                "paper_set": seat.get("paper_set"),
            })
    return jsonify(results[:10])

@app.route("/api/export/<exam_id>")
@login_required
def export_csv(exam_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM seating WHERE exam_id=? ORDER BY created_at DESC LIMIT 1",(exam_id,)).fetchone()
    conn.close()
    if not row: return jsonify({"error":"No arrangement found"}),404
    arr = json.loads(row["data"]); seats = arr.get("seats",[])
    out = io.StringIO(); w = csv.writer(out)
    w.writerow(["Seat No.","Grid Position","Student Name","Register Number","Department","Subject","Room","Paper Set"])
    for s in seats:
        w.writerow([s.get("seat_number"),s.get("seat_label"),s.get("student_name"),
                    s.get("register_number"),s.get("department"),s.get("subject"),
                    s.get("room_name"),s.get("paper_set")])
    out.seek(0)
    name = arr.get("exam",{}).get("name","seating")
    return send_file(io.BytesIO(out.read().encode()), mimetype="text/csv",
                     as_attachment=True, download_name=f"{name}_seating.csv")

@app.route("/api/ai-analyze", methods=["POST"])
@login_required
def ai_analyze():
    exam_id = request.json.get("exam_id")
    conn = get_db()
    row = conn.execute("SELECT * FROM seating WHERE exam_id=? ORDER BY created_at DESC LIMIT 1",(exam_id,)).fetchone()
    conn.close()
    if not row: return jsonify({"error":"No arrangement to analyze"}),404
    arr = json.loads(row["data"]); seats = arr.get("seats",[])
    dept_dist = {}
    for s in seats:
        room = s.get("room_name","?"); dept = s.get("department","?")
        dept_dist.setdefault(room,{}); dept_dist[room][dept] = dept_dist[room].get(dept,0)+1
    summary = f"""Analyze exam seating for malpractice risk:
- Students: {len(seats)}, Rooms: {len(set(s['room_id'] for s in seats))}
- Departments: {list(set(s.get('department','') for s in seats))}
- Distribution: {json.dumps(dept_dist)}
- Settings: {json.dumps(arr.get('settings',{}))}
Reply ONLY with JSON: risk_score(0-10), issues(list), suggestions(list), quality_score(0-10), summary(string)"""
    api_key = os.getenv("GROK_API_KEY")
    if not api_key or api_key == "your_grok_api_key_here":
        return jsonify({"error":"GROK_API_KEY not set in .env"}),400
    try:
        client = OpenAI(api_key=api_key, base_url="https://api.x.ai/v1")
        msg = client.chat.completions.create(model="grok-3-fast-beta", max_tokens=1024,
            messages=[{"role":"system","content":"Reply only with valid JSON."},
                      {"role":"user","content":summary}])
        text = msg.choices[0].message.content.strip()
        if text.startswith("```"): text = re.sub(r"```[a-z]*\n?","",text).replace("```","").strip()
        try: analysis = json.loads(text)
        except:
            m = re.search(r'\{.*\}', text, re.DOTALL)
            analysis = json.loads(m.group()) if m else {"summary":text}
        return jsonify({"success":True,"analysis":analysis})
    except Exception as e:
        return jsonify({"error":f"Grok API error: {str(e)}"}),500

# ── PRINT EXPORT ──────────────────────────────────────────────────────────────
@app.route("/api/print/<exam_id>/<room_id>")
@login_required
def print_hall(exam_id, room_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM seating WHERE exam_id=? ORDER BY created_at DESC LIMIT 1",(exam_id,)).fetchone()
    conn.close()
    if not row: return "No seating arrangement found", 404

    arr = json.loads(row["data"])
    exam = arr.get("exam", {})
    all_seats = arr.get("seats", [])
    rooms_meta = {r["id"]: r for r in arr.get("rooms", [])}
    room_seats = [s for s in all_seats if s.get("room_id") == room_id]
    if not room_seats: return "No seats found for this hall", 404

    room_name = room_seats[0].get("room_name", "Hall")
    room_meta = rooms_meta.get(room_id, {})
    max_row = max(s.get("row", 1) for s in room_seats)
    max_col = max(s.get("col", 1) for s in room_seats)

    # Invigilator assignment
    invigilators = arr.get("invigilators", [])
    room_ids_list = list(dict.fromkeys(s.get("room_id") for s in all_seats))
    room_idx = room_ids_list.index(room_id) if room_id in room_ids_list else 0
    invigilator = invigilators[room_idx % len(invigilators)]["name"] if invigilators else "—"

    # Parse blocked seats
    blocked = room_meta.get("blocked_seats", [])
    if isinstance(blocked, str):
        try: blocked = json.loads(blocked)
        except: blocked = []

    # Parse aisles
    def parse_aisles_print(raw):
        if not raw: return []
        if isinstance(raw, list): return [int(x) for x in raw if str(x).strip().isdigit() or isinstance(x,int)]
        raw = str(raw).strip()
        if raw.startswith("["): 
            try: return [int(x) for x in json.loads(raw)]
            except: return []
        return [int(x.strip()) for x in raw.split(",") if x.strip().isdigit()]

    aisles = parse_aisles_print(room_meta.get("aisle_after_col"))
    if not aisles and room_seats and room_seats[0].get("aisles"):
        aisles = parse_aisles_print(room_seats[0]["aisles"])

    # Build seat map
    seat_map = {(s["row"], s["col"]): s for s in room_seats}

    # Build grid for template
    grid = []
    for r in range(1, max_row + 1):
        row_cells = []
        for c in range(1, max_col + 1):
            lbl = f"R{r}C{c}"
            seat = seat_map.get((r, c))
            if lbl in blocked:
                row_cells.append({"type": "blocked"})
            elif seat:
                row_cells.append({
                    "type": "seat",
                    "seat_number": seat.get("seat_number", ""),
                    "register_number": seat.get("register_number", ""),
                    "department": seat.get("department", ""),
                    "year": seat.get("year", ""),
                    "subject": seat.get("subject", ""),
                    "paper_set": seat.get("paper_set", "A"),
                })
            else:
                row_cells.append({"type": "empty"})
            # Insert aisle gap after this column
            if c in aisles:
                row_cells.append({"type": "aisle"})
        grid.append(row_cells)

    # Department summary
    from collections import Counter as _C
    summary_data = _C(f"{s.get('department','')} Y{s.get('year','')}" for s in room_seats)
    summary = sorted(summary_data.items())

    # Format time
    raw_time = exam.get("time", "")
    nice_time = ""
    if raw_time:
        try:
            h, m = raw_time.split(":")
            h = int(h)
            nice_time = f"{h if h<=12 else h-12}:{m} {'AM' if h<12 else 'PM'}"
        except: nice_time = raw_time

    # Dept/Subject string
    unique_ds = []
    for s in room_seats:
        dept = s.get("department", "").strip()
        yr = s.get("year", "").strip()
        dy = f"{dept} Y{yr}" if yr else dept
        subj = s.get("subject", "").strip()
        lbl = f"{dy}: {subj}" if subj else dy
        if lbl not in unique_ds:
            unique_ds.append(lbl)
    
    display_subject = " | ".join(unique_ds) if unique_ds else (exam.get("subject", "") or "—")

    html = render_template("print_hall.html",
        room_name=room_name,
        exam_name=exam.get("name", "Exam"),
        subject=display_subject,
        subject_code=exam.get("subject_code", ""),
        exam_date=exam.get("date", ""),
        exam_time=nice_time,
        duration=exam.get("duration", ""),
        invigilator=invigilator,
        total_students=len(room_seats),
        summary=summary,
        grid=grid,
        seat_min=min(s.get("seat_number", 0) for s in room_seats),
        seat_max=max(s.get("seat_number", 0) for s in room_seats),
        generated_at=datetime.now().strftime("%d %b %Y %I:%M %p"),
    )
    resp = make_response(html)
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp

@app.route("/api/print/<exam_id>")
@login_required
def print_all_halls(exam_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM seating WHERE exam_id=? ORDER BY created_at DESC LIMIT 1",(exam_id,)).fetchone()
    conn.close()
    if not row: return "No seating found", 404
    arr = json.loads(row["data"])
    room_ids = list(dict.fromkeys(s.get("room_id") for s in arr.get("seats", [])))
    pages = []
    for rid in room_ids:
        result = print_hall(exam_id, rid)
        page_html = result.get_data(as_text=True)
        body_match = re.search(r'<body[^>]*>(.*?)</body>', page_html, re.DOTALL)
        if body_match:
            pages.append(body_match.group(1))
    combined = render_template("print_all_halls.html",
        exam_id=exam_id,
        pages=pages
    )
    resp = make_response(combined)
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    return resp

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=os.getenv("FLASK_DEBUG","0")=="1", host="0.0.0.0", port=5000)
